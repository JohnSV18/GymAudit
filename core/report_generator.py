"""
Report Generator Module
Creates Excel audit reports with highlighting and formatting
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List, Dict, Any
from pathlib import Path


class AuditReportGenerator:
    """Generates formatted Excel audit reports"""

    # Styling
    HIGHLIGHT_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow (flagged)
    BP_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange (BP accounts)
    XX_FILL = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')  # Light blue (XX code accounts)
    HEADER_FILL = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Gray (column headers)
    SECTION_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue (section headers)
    BOLD_FONT = Font(bold=True)
    WHITE_FONT = Font(bold=True, color='FFFFFF')
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    MEMBER_SEPARATOR = Border(bottom=Side(style='thick'))

    def _apply_member_separator(self, sheet, row_num: int, col_count: int):
        """Apply thick bottom border to all cells in a row to separate member groups."""
        for col_idx in range(1, col_count + 1):
            cell = sheet.cell(row=row_num, column=col_idx)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=Side(style='thick')
            )

    def __init__(self, output_folder: str = 'outputs'):
        """
        Initialize report generator

        Args:
            output_folder: Directory to save reports
        """
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)

    def _is_bp_member(self, row: List[str], column_indices: List[int], bp_config: Dict[str, Any] = None) -> bool:
        """
        Check if row has BP/Billing indicators based on config.

        Args:
            row: Data row
            column_indices: List of column indices to check
            bp_config: BP detection configuration with 'keywords' and 'case_sensitive'

        Returns:
            True if member has BP/Billing indicator
        """
        if bp_config and not bp_config.get('enabled', True):
            return False

        keywords = ['bp', 'billing']
        case_sensitive = False

        if bp_config:
            keywords = bp_config.get('keywords', keywords)
            case_sensitive = bp_config.get('case_sensitive', False)

        for idx in column_indices:
            if idx is not None and idx >= 0 and idx < len(row):
                value = str(row[idx])
                if not case_sensitive:
                    value = value.lower()
                    check_keywords = [k.lower() for k in keywords]
                else:
                    check_keywords = keywords

                for keyword in check_keywords:
                    if keyword in value:
                        return True
        return False

    def _is_xx_code(self, row: List[str], code_column_index: int) -> bool:
        """
        Check if a row has 'xx' in the code column.

        Args:
            row: Data row
            code_column_index: Index of the code column

        Returns:
            True if code column value is 'xx' (case-insensitive)
        """
        if code_column_index is not None and 0 <= code_column_index < len(row):
            value = str(row[code_column_index]).strip().lower()
            if value == 'xx':
                return True
        return False

    def _get_code_column_index(self, column_mapping: Dict[str, int] = None) -> int:
        """Get the code column index from column_mapping or use default."""
        if column_mapping:
            idx = column_mapping.get('code')
            if idx is not None and idx >= 0:
                return idx
        return 11  # default new format code column

    def _categorize_members(
        self,
        member_results: Dict[str, Any],
        column_indices: List[int],
        bp_config: Dict[str, Any],
        code_column_index: int = None
    ) -> tuple:
        """
        Categorize grouped members into flagged, XX, BP, and valid lists.

        Args:
            member_results: Dict of member_number -> member result data
            column_indices: Column indices for BP detection
            bp_config: BP detection configuration
            code_column_index: Index of code column for XX detection

        Returns:
            Tuple of (flagged, xx, bp, valid) member result lists
        """
        flagged, xx, bp, valid = [], [], [], []

        for member_number, result in member_results.items():
            is_bp = False
            is_xx = False
            for txn in result.get('transactions', []):
                if self._is_bp_member(txn, column_indices, bp_config):
                    is_bp = True
                    break
                if code_column_index is not None and self._is_xx_code(txn, code_column_index):
                    is_xx = True

            if is_bp:
                bp.append(result)
            elif is_xx:
                xx.append(result)
            elif result.get('has_flags', False):
                flagged.append(result)
            else:
                valid.append(result)

        return flagged, xx, bp, valid

    def _categorize_rows(
        self,
        data_rows: List[List[str]],
        audit_results: List[Dict[str, Any]],
        column_indices: List[int],
        bp_config: Dict[str, Any] = None,
        code_column_index: int = None
    ) -> tuple:
        """
        Split rows into flagged, XX, BP, and valid categories.

        Args:
            data_rows: Original data rows
            audit_results: Audit results for each row
            column_indices: List of column indices to check for BP detection
            bp_config: BP detection configuration
            code_column_index: Index of code column for XX detection

        Returns:
            Tuple of (flagged, xx, bp, valid) lists, each containing (row, result) tuples
        """
        flagged, xx, bp, valid = [], [], [], []

        for row, result in zip(data_rows, audit_results):
            is_bp = self._is_bp_member(row, column_indices, bp_config)
            is_xx = self._is_xx_code(row, code_column_index) if code_column_index is not None else False
            has_flags = bool(result.get('red_flags'))

            if is_bp:
                bp.append((row, result))
            elif is_xx:
                xx.append((row, result))
            elif has_flags:
                flagged.append((row, result))
            else:
                valid.append((row, result))

        return flagged, xx, bp, valid

    def _write_section_header(self, sheet, row_num: int, header_text: str, col_count: int) -> int:
        """
        Write a section header row.

        Args:
            sheet: Excel worksheet
            row_num: Row number to write at
            header_text: Text for the section header
            col_count: Number of columns to merge

        Returns:
            Next available row number
        """
        cell = sheet.cell(row=row_num, column=1, value=header_text)
        cell.font = self.WHITE_FONT
        cell.fill = self.SECTION_FILL
        sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
        return row_num + 1

    def create_audit_report(
        self,
        header_row: List[str],
        data_rows: List[List[str]],
        audit_results: List[Dict[str, Any]],
        output_filename: str,
        include_summary_sheet: bool = True,
        column_mapping: Dict[str, int] = None,
        bp_config: Dict[str, Any] = None
    ) -> str:
        """
        Create Excel audit report with highlighted red flags organized into sections.

        Sections:
        1. FLAGGED ACCOUNTS - Members with red flags (excluding BP members)
        2. BILLING PROBLEM ACCOUNTS - Members with BP/Billing codes
        3. VALID ACCOUNTS - Members with no flags and no BP codes

        Args:
            header_row: Column headers
            data_rows: Original data rows
            audit_results: List of audit results for each row (from audit_engine)
            output_filename: Name for output file
            include_summary_sheet: Whether to add a summary sheet
            column_mapping: Dictionary with column name to index mappings
            bp_config: BP detection configuration with 'enabled', 'columns', 'keywords', 'case_sensitive'

        Returns:
            Full path to generated report file
        """
        wb = Workbook()

        # Create main audit sheet
        audit_sheet = wb.active
        audit_sheet.title = "Audit Report"

        # Add only "Notes" column to header (keep original columns intact)
        enhanced_header = header_row + ["Notes"]
        col_count = len(enhanced_header)

        # Write header row with formatting
        for col_idx, header_text in enumerate(enhanced_header, start=1):
            cell = audit_sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN

        # Get column indices for BP detection
        column_indices = []
        if column_mapping:
            # Use column_mapping to get indices for configured columns
            bp_columns = bp_config.get('columns', ['code', 'member_type']) if bp_config else ['code', 'member_type']
            for col_name in bp_columns:
                idx = column_mapping.get(col_name)
                if idx is not None and idx >= 0:
                    column_indices.append(idx)
        if not column_indices:
            # Default to old format indices for code (7) and member_type (5)
            column_indices = [7, 5]

        # Get code column index for XX detection
        code_col_idx = self._get_code_column_index(column_mapping)

        # Categorize rows into sections
        flagged_rows, xx_rows, bp_rows, valid_rows = self._categorize_rows(
            data_rows, audit_results, column_indices, bp_config, code_col_idx
        )

        excel_row = 2
        flagged_count = 0
        bp_count = 0
        xx_count = 0

        # Section 1: FLAGGED ACCOUNTS
        if flagged_rows:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"FLAGGED ACCOUNTS ({len(flagged_rows)} records)",
                col_count
            )

            for data_row, audit_result in flagged_rows:
                red_flags = audit_result.get('red_flags', [])
                notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                enhanced_row = data_row + [notes]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.HIGHLIGHT_FILL  # Yellow

                flagged_count += 1
                excel_row += 1

            excel_row += 1

        # Section 2: XX CODE ACCOUNTS
        if xx_rows:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"XX CODE ACCOUNTS ({len(xx_rows)} records)",
                col_count
            )

            for data_row, audit_result in xx_rows:
                red_flags = audit_result.get('red_flags', [])
                notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                enhanced_row = data_row + [notes]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.XX_FILL

                xx_count += 1
                excel_row += 1

            excel_row += 1

        # Section 3: BILLING PROBLEM ACCOUNTS
        if bp_rows:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"BILLING PROBLEM ACCOUNTS ({len(bp_rows)} records)",
                col_count
            )

            for data_row, audit_result in bp_rows:
                red_flags = audit_result.get('red_flags', [])
                notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                enhanced_row = data_row + [notes]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.BP_FILL  # Orange

                bp_count += 1
                excel_row += 1

            excel_row += 1

        # Section 4: VALID ACCOUNTS
        if valid_rows:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"VALID ACCOUNTS ({len(valid_rows)} records)",
                col_count
            )

            for data_row, audit_result in valid_rows:
                enhanced_row = data_row + [""]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)

                excel_row += 1

        # Auto-adjust column widths
        self._auto_adjust_column_widths(audit_sheet)

        # Add summary sheet if requested
        if include_summary_sheet:
            self._add_summary_sheet(wb, audit_results, flagged_count, len(data_rows), bp_count, xx_count)

        # Save workbook
        output_path = self.output_folder / output_filename
        wb.save(output_path)

        return str(output_path)

    def create_grouped_audit_report(
        self,
        header_row: List[str],
        grouped_results: Dict[str, Any],
        output_filename: str,
        column_mapping: Dict[str, int] = None,
        bp_config: Dict[str, Any] = None
    ) -> str:
        """
        Create Excel audit report for grouped member-based audits.
        Rows are grouped by member_number with a Net Balance column.

        Sections:
        1. FLAGGED ACCOUNTS - Members with red flags
        2. BILLING PROBLEM ACCOUNTS - Members with BP/Billing codes
        3. VALID ACCOUNTS - Members with no flags

        Args:
            header_row: Column headers from original file
            grouped_results: Results from audit_pif_grouped()
            output_filename: Name for output file
            column_mapping: Dictionary with column name to index mappings
            bp_config: BP detection configuration

        Returns:
            Full path to generated report file
        """
        wb = Workbook()
        audit_sheet = wb.active
        audit_sheet.title = "Audit Report"

        # Add "Net Balance" and "Notes" columns to header
        enhanced_header = header_row + ["Net Balance", "Notes"]
        col_count = len(enhanced_header)

        # Write header row with formatting
        for col_idx, header_text in enumerate(enhanced_header, start=1):
            cell = audit_sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN

        # Get column indices for BP detection
        column_indices = []
        if column_mapping:
            bp_columns = bp_config.get('columns', ['code', 'member_type']) if bp_config else ['code', 'member_type']
            for col_name in bp_columns:
                idx = column_mapping.get(col_name)
                if idx is not None and idx >= 0:
                    column_indices.append(idx)
        if not column_indices:
            column_indices = [11, 9, 10]

        # Get code column index for XX detection
        code_col_idx = self._get_code_column_index(column_mapping)

        # Categorize members into flagged, XX, BP, and valid
        member_results = grouped_results.get('member_results', {})
        flagged_members, xx_members, bp_members, valid_members = self._categorize_members(
            member_results, column_indices, bp_config, code_col_idx
        )

        excel_row = 2
        flagged_count = 0
        bp_count = 0
        xx_count = 0

        # Section 1: FLAGGED ACCOUNTS
        if flagged_members:
            total_flagged_rows = sum(len(m['transactions']) for m in flagged_members)
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"FLAGGED ACCOUNTS ({len(flagged_members)} members, {total_flagged_rows} records)",
                col_count
            )

            for member_result in flagged_members:
                flags = member_result.get('flags', [])
                notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                        cell.fill = self.HIGHLIGHT_FILL
                    excel_row += 1

                self._apply_member_separator(audit_sheet, excel_row - 1, col_count)
                flagged_count += 1

            excel_row += 1

        # Section 2: XX CODE ACCOUNTS
        if xx_members:
            total_xx_rows = sum(len(m['transactions']) for m in xx_members)
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"XX CODE ACCOUNTS ({len(xx_members)} members, {total_xx_rows} records)",
                col_count
            )

            for member_result in xx_members:
                flags = member_result.get('flags', [])
                notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                        cell.fill = self.XX_FILL
                    excel_row += 1

                self._apply_member_separator(audit_sheet, excel_row - 1, col_count)
                xx_count += 1

            excel_row += 1

        # Section 3: BILLING PROBLEM ACCOUNTS
        if bp_members:
            total_bp_rows = sum(len(m['transactions']) for m in bp_members)
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"BILLING PROBLEM ACCOUNTS ({len(bp_members)} members, {total_bp_rows} records)",
                col_count
            )

            for member_result in bp_members:
                flags = member_result.get('flags', [])
                notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                        cell.fill = self.BP_FILL
                    excel_row += 1

                self._apply_member_separator(audit_sheet, excel_row - 1, col_count)
                bp_count += 1

            excel_row += 1

        # Section 4: VALID ACCOUNTS
        if valid_members:
            total_valid_rows = sum(len(m['transactions']) for m in valid_members)
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"VALID ACCOUNTS ({len(valid_members)} members, {total_valid_rows} records)",
                col_count
            )

            for member_result in valid_members:
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", ""]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                    excel_row += 1

                self._apply_member_separator(audit_sheet, excel_row - 1, col_count)

        # Auto-adjust column widths
        self._auto_adjust_column_widths(audit_sheet)

        # Build flat audit_results for summary sheet
        audit_results = []
        for member_result in member_results.values():
            audit_results.append({
                'red_flags': member_result.get('flags', []),
                'has_flags': member_result.get('has_flags', False),
                'financial_impact': member_result.get('net_balance', 0) if member_result.get('net_balance', 0) > 0 else 0,
                'dues_impact': 0,
                'balance_impact': member_result.get('net_balance', 0) if member_result.get('net_balance', 0) > 0 else 0
            })

        # Add summary sheet
        total_count = sum(len(m['transactions']) for m in member_results.values())
        self._add_summary_sheet(wb, audit_results, flagged_count, total_count, bp_count, xx_count)

        # Save workbook
        output_path = self.output_folder / output_filename
        wb.save(output_path)

        return str(output_path)

    def _auto_adjust_column_widths(self, sheet):
        """Auto-adjust column widths based on content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set width (max 50 for readability)
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _add_summary_sheet(self, workbook, audit_results: List[Dict[str, Any]], flagged_count: int, total_count: int, bp_count: int = 0, xx_count: int = 0):
        """Add a summary sheet with statistics"""
        summary_sheet = workbook.create_sheet("Summary", 0)  # Insert at beginning

        # Title
        summary_sheet['A1'] = "AUDIT SUMMARY"
        summary_sheet['A1'].font = Font(bold=True, size=16)

        # Overall stats
        row = 3
        summary_sheet[f'A{row}'] = "Total Records:"
        summary_sheet[f'B{row}'] = total_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        row += 1
        summary_sheet[f'A{row}'] = "Flagged Records:"
        summary_sheet[f'B{row}'] = flagged_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].fill = self.HIGHLIGHT_FILL

        row += 1
        summary_sheet[f'A{row}'] = "XX Code Records:"
        summary_sheet[f'B{row}'] = xx_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].fill = self.XX_FILL

        row += 1
        summary_sheet[f'A{row}'] = "Billing Problem Records:"
        summary_sheet[f'B{row}'] = bp_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].fill = self.BP_FILL

        row += 1
        valid_count = total_count - flagged_count - bp_count - xx_count
        summary_sheet[f'A{row}'] = "Valid Records:"
        summary_sheet[f'B{row}'] = valid_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        row += 1
        flagged_percentage = (flagged_count / total_count * 100) if total_count > 0 else 0
        summary_sheet[f'A{row}'] = "Flagged Percentage:"
        summary_sheet[f'B{row}'] = f"{flagged_percentage:.1f}%"
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        # Red flag breakdown
        row += 3
        summary_sheet[f'A{row}'] = "RED FLAG BREAKDOWN"
        summary_sheet[f'A{row}'].font = Font(bold=True, size=14)

        # Count red flags by type
        flag_counts = {}
        for result in audit_results:
            for flag in result.get('red_flags', []):
                flag_type = flag.flag_type
                flag_counts[flag_type] = flag_counts.get(flag_type, 0) + 1

        # Sort by count (descending)
        sorted_flags = sorted(flag_counts.items(), key=lambda x: x[1], reverse=True)

        row += 2
        summary_sheet[f'A{row}'] = "Red Flag Type"
        summary_sheet[f'B{row}'] = "Count"
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].font = self.BOLD_FONT

        for flag_type, count in sorted_flags:
            row += 1
            summary_sheet[f'A{row}'] = self._format_flag_type(flag_type)
            summary_sheet[f'B{row}'] = count

        # Financial impact
        row += 3
        summary_sheet[f'A{row}'] = "FINANCIAL IMPACT"
        summary_sheet[f'A{row}'].font = Font(bold=True, size=14)

        total_financial_impact = sum(result.get('financial_impact', 0) for result in audit_results)

        row += 2
        summary_sheet[f'A{row}'] = "Total Potential Revenue at Risk:"
        summary_sheet[f'B{row}'] = f"${total_financial_impact:,.2f}"
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].font = Font(bold=True, color="FF0000")

        # Auto-adjust widths
        summary_sheet.column_dimensions['A'].width = 35
        summary_sheet.column_dimensions['B'].width = 20

    def _format_flag_type(self, flag_type: str) -> str:
        """Format flag type for display"""
        # Convert snake_case to Title Case
        formatted = flag_type.replace('_', ' ').title()

        # Specific formatting
        replacements = {
            'Dues Low': 'Dues Below Minimum',
            'Dues Invalid': 'Invalid Dues Amount',
            'Date Mismatch': 'Join/Exp Date Mismatch',
            'Date Invalid': 'Invalid Date Format',
            'Pay Type Wrong': 'Incorrect Pay Type',
            'End Draft Wrong': 'Incorrect End Draft Date',
            'Cycle Wrong': 'Incorrect Cycle Number',
            'Cycle Invalid': 'Invalid Cycle Value',
            'Balance Debit': 'Outstanding Balance (Owed)',
            'Balance Credit': 'Credit Balance (Refund Due)',
            'Balance Invalid': 'Invalid Balance Value',
            # Grouped matching flags
            'Unpaid Balance': 'Unpaid Balance (Charge Without Payment)',
            'Overpayment': 'Overpayment (Payment Exceeds Charges)',
            'Member Name Mismatch': 'Member Name Mismatch',
            'Price Mismatch': 'Price Mismatch (Wrong Amount)',
            'Low Amount': 'Low Transaction Amount',
            # MTM-specific flags
            'Needs Verification': 'Charge Without Matching Payment',
            'Missing Monthly Payment': 'Missing Monthly Payment',
            'Missing Enrollment Fee': 'Missing Enrollment Fee ($50)',
            'Missing Annual Fee': 'Missing Annual Fee',
            'Exp Year Wrong': 'Expiration Year Not 2099',
            'End Draft Year Wrong': 'End Draft Year Not 2099',
            'Draft Date Too Far': 'Draft Date Too Far From Join'
        }

        return replacements.get(formatted, formatted)

    def create_consolidated_report(
        self,
        file_results: List[Dict[str, Any]],
        output_filename: str = "Consolidated_Audit_Report.xlsx"
    ) -> str:
        """
        Create a consolidated report across multiple files

        Args:
            file_results: List of file audit results
            output_filename: Name for output file

        Returns:
            Full path to generated report file
        """
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Overview"

        # Title
        summary_sheet['A1'] = "CONSOLIDATED AUDIT REPORT"
        summary_sheet['A1'].font = Font(bold=True, size=16)

        # Per-file summary
        row = 3
        summary_sheet['A3'] = "Filename"
        summary_sheet['B3'] = "Total Records"
        summary_sheet['C3'] = "Flagged"
        summary_sheet['D3'] = "Clean"
        summary_sheet['E3'] = "Flag %"
        summary_sheet['F3'] = "Financial Impact"

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            summary_sheet[f'{col}3'].font = self.BOLD_FONT
            summary_sheet[f'{col}3'].fill = self.HEADER_FILL

        row = 4
        total_records = 0
        total_flagged = 0
        total_impact = 0

        for file_result in file_results:
            filename = file_result['filename']
            records = file_result['total_records']
            flagged = file_result['flagged_count']
            clean = records - flagged
            flag_pct = (flagged / records * 100) if records > 0 else 0
            impact = file_result.get('total_financial_impact', 0)

            summary_sheet[f'A{row}'] = filename
            summary_sheet[f'B{row}'] = records
            summary_sheet[f'C{row}'] = flagged
            summary_sheet[f'D{row}'] = clean
            summary_sheet[f'E{row}'] = f"{flag_pct:.1f}%"
            summary_sheet[f'F{row}'] = f"${impact:,.2f}"

            if flagged > 0:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    summary_sheet[f'{col}{row}'].fill = self.HIGHLIGHT_FILL

            total_records += records
            total_flagged += flagged
            total_impact += impact

            row += 1

        # Totals row
        row += 1
        summary_sheet[f'A{row}'] = "TOTALS"
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'] = total_records
        summary_sheet[f'C{row}'] = total_flagged
        summary_sheet[f'D{row}'] = total_records - total_flagged
        total_flag_pct = (total_flagged / total_records * 100) if total_records > 0 else 0
        summary_sheet[f'E{row}'] = f"{total_flag_pct:.1f}%"
        summary_sheet[f'F{row}'] = f"${total_impact:,.2f}"

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            summary_sheet[f'{col}{row}'].font = self.BOLD_FONT

        # Auto-adjust widths
        summary_sheet.column_dimensions['A'].width = 40
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 15
        summary_sheet.column_dimensions['E'].width = 12
        summary_sheet.column_dimensions['F'].width = 20

        # Save
        output_path = self.output_folder / output_filename
        wb.save(output_path)

        return str(output_path)

    def create_mtm_audit_report(
        self,
        header_row: List[str],
        mtm_results: Dict[str, Any],
        output_filename: str,
        column_mapping: Dict[str, int] = None,
        bp_config: Dict[str, Any] = None
    ) -> str:
        """
        Create Excel audit report for Month-to-Month transaction audits.

        Organized by member with sections:
        1. FLAGGED MEMBERS - Members with missing payments or other flags
        2. BILLING PROBLEM MEMBERS - Members with BP indicators
        3. VALID MEMBERS - Members with no issues

        Args:
            header_row: Column headers from original file
            mtm_results: Results from audit_month_to_month_transactions()
            output_filename: Name for output file
            column_mapping: Dictionary with column name to index mappings
            bp_config: BP detection configuration

        Returns:
            Full path to generated report file
        """
        wb = Workbook()

        # Create main audit sheet
        audit_sheet = wb.active
        audit_sheet.title = "MTM Audit Report"

        # Create member-level header (different from transaction header)
        member_header = [
            "First Name", "Last Name", "Member #", "Join Date",
            "Member Type", "Enrollment Fee", "Initial Payment", "Coverage Start",
            "Months Paid", "Missing Months", "Annual Fee", "Unmatched Charges", "Notes"
        ]
        col_count = len(member_header)

        # Write header row
        for col_idx, header_text in enumerate(member_header, start=1):
            cell = audit_sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN

        # Get column indices for BP detection
        column_indices = []
        if column_mapping:
            bp_columns = bp_config.get('columns', ['code', 'member_type', 'member_group']) if bp_config else ['code', 'member_type', 'member_group']
            for col_name in bp_columns:
                idx = column_mapping.get(col_name)
                if idx is not None and idx >= 0:
                    column_indices.append(idx)

        # Get code column index for XX detection
        code_col_idx = self._get_code_column_index(column_mapping)

        # Categorize members into flagged, XX, BP, and valid
        member_results = mtm_results.get('member_results', {})
        flagged_members, xx_members, bp_members, valid_members = self._categorize_members(
            member_results, column_indices, bp_config, code_col_idx
        )

        excel_row = 2
        flagged_count = 0
        bp_count = 0
        xx_count = 0

        # Section 1: FLAGGED MEMBERS
        if flagged_members:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"FLAGGED MEMBERS ({len(flagged_members)} members)",
                col_count
            )

            for result in flagged_members:
                row_data = self._build_mtm_row_data(result)

                for col_idx, value in enumerate(row_data, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.HIGHLIGHT_FILL

                self._apply_member_separator(audit_sheet, excel_row, col_count)
                flagged_count += 1
                excel_row += 1

            excel_row += 1

        # Section 2: XX CODE MEMBERS
        if xx_members:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"XX CODE MEMBERS ({len(xx_members)} members)",
                col_count
            )

            for result in xx_members:
                row_data = self._build_mtm_row_data(result)

                for col_idx, value in enumerate(row_data, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.XX_FILL

                self._apply_member_separator(audit_sheet, excel_row, col_count)
                xx_count += 1
                excel_row += 1

            excel_row += 1

        # Section 3: BILLING PROBLEM MEMBERS
        if bp_members:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"BILLING PROBLEM MEMBERS ({len(bp_members)} members)",
                col_count
            )

            for result in bp_members:
                row_data = self._build_mtm_row_data(result)

                for col_idx, value in enumerate(row_data, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.BP_FILL

                self._apply_member_separator(audit_sheet, excel_row, col_count)
                bp_count += 1
                excel_row += 1

            excel_row += 1

        # Section 4: VALID MEMBERS
        if valid_members:
            excel_row = self._write_section_header(
                audit_sheet, excel_row,
                f"VALID MEMBERS ({len(valid_members)} members)",
                col_count
            )

            for result in valid_members:
                row_data = self._build_mtm_row_data(result)

                for col_idx, value in enumerate(row_data, start=1):
                    cell = audit_sheet.cell(row=excel_row, column=col_idx, value=value)

                self._apply_member_separator(audit_sheet, excel_row, col_count)
                excel_row += 1

        # Auto-adjust column widths
        self._auto_adjust_column_widths(audit_sheet)

        # Add summary sheet
        self._add_mtm_summary_sheet(
            wb, mtm_results, flagged_count, bp_count, len(valid_members), xx_count
        )

        # Add transactions detail sheet
        self._add_transactions_sheet(wb, header_row, mtm_results, column_indices, bp_config)

        # Save workbook
        output_path = self.output_folder / output_filename
        wb.save(output_path)

        return str(output_path)

    def _add_mtm_summary_sheet(
        self,
        workbook,
        mtm_results: Dict[str, Any],
        flagged_count: int,
        bp_count: int,
        valid_count: int,
        xx_count: int = 0
    ):
        """Add summary sheet for MTM audit"""
        summary_sheet = workbook.create_sheet("Summary", 0)

        # Title
        summary_sheet['A1'] = "MTM AUDIT SUMMARY"
        summary_sheet['A1'].font = Font(bold=True, size=16)

        # Overall stats
        row = 3
        total_members = mtm_results.get('total_members', 0)
        total_transactions = mtm_results.get('total_transactions', 0)

        summary_sheet[f'A{row}'] = "Total Members:"
        summary_sheet[f'B{row}'] = total_members
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        row += 1
        summary_sheet[f'A{row}'] = "Total Transactions:"
        summary_sheet[f'B{row}'] = total_transactions
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        row += 2
        summary_sheet[f'A{row}'] = "Flagged Members:"
        summary_sheet[f'B{row}'] = flagged_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].fill = self.HIGHLIGHT_FILL

        row += 1
        summary_sheet[f'A{row}'] = "XX Code Members:"
        summary_sheet[f'B{row}'] = xx_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].fill = self.XX_FILL

        row += 1
        summary_sheet[f'A{row}'] = "Billing Problem Members:"
        summary_sheet[f'B{row}'] = bp_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].fill = self.BP_FILL

        row += 1
        summary_sheet[f'A{row}'] = "Valid Members:"
        summary_sheet[f'B{row}'] = valid_count
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        row += 1
        flagged_percentage = (flagged_count / total_members * 100) if total_members > 0 else 0
        summary_sheet[f'A{row}'] = "Flagged Percentage:"
        summary_sheet[f'B{row}'] = f"{flagged_percentage:.1f}%"
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        # Member type breakdown
        row += 3
        summary_sheet[f'A{row}'] = "MEMBER TYPE BREAKDOWN"
        summary_sheet[f'A{row}'].font = Font(bold=True, size=14)

        member_results = mtm_results.get('member_results', {})
        new_members = sum(1 for r in member_results.values() if r.get('member_type') == 'New')
        existing_members = sum(1 for r in member_results.values() if r.get('member_type') == 'Existing')
        with_enrollment = sum(1 for r in member_results.values() if r.get('has_enrollment_fee'))
        with_initial = sum(1 for r in member_results.values() if r.get('has_initial_payment'))
        with_annual = sum(1 for r in member_results.values() if r.get('has_annual_fee'))

        row += 2
        summary_sheet[f'A{row}'] = "New Members (with enrollment fee):"
        summary_sheet[f'B{row}'] = new_members
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        row += 1
        summary_sheet[f'A{row}'] = "Existing Members (no enrollment fee):"
        summary_sheet[f'B{row}'] = existing_members
        summary_sheet[f'A{row}'].font = self.BOLD_FONT

        row += 2
        summary_sheet[f'A{row}'] = "Members with Initial Payment:"
        summary_sheet[f'B{row}'] = with_initial

        row += 1
        summary_sheet[f'A{row}'] = "Members with Annual Fee:"
        summary_sheet[f'B{row}'] = with_annual

        # Red flag breakdown
        row += 3
        summary_sheet[f'A{row}'] = "RED FLAG BREAKDOWN"
        summary_sheet[f'A{row}'].font = Font(bold=True, size=14)

        # Count red flags by type
        flag_counts = {}
        member_results = mtm_results.get('member_results', {})
        for result in member_results.values():
            for flag in result.get('flags', []):
                flag_type = flag.flag_type
                flag_counts[flag_type] = flag_counts.get(flag_type, 0) + 1

        sorted_flags = sorted(flag_counts.items(), key=lambda x: x[1], reverse=True)

        row += 2
        summary_sheet[f'A{row}'] = "Red Flag Type"
        summary_sheet[f'B{row}'] = "Count"
        summary_sheet[f'A{row}'].font = self.BOLD_FONT
        summary_sheet[f'B{row}'].font = self.BOLD_FONT

        for flag_type, count in sorted_flags:
            row += 1
            summary_sheet[f'A{row}'] = self._format_flag_type(flag_type)
            summary_sheet[f'B{row}'] = count

        # Auto-adjust widths
        summary_sheet.column_dimensions['A'].width = 35
        summary_sheet.column_dimensions['B'].width = 20

    def _add_transactions_sheet(
        self,
        workbook,
        header_row: List[str],
        mtm_results: Dict[str, Any],
        column_indices: List[int],
        bp_config: Dict[str, Any]
    ):
        """Add detailed transactions sheet"""
        txn_sheet = workbook.create_sheet("All Transactions")

        # Write header
        enhanced_header = header_row + ["Member Status", "Flags"]
        for col_idx, header_text in enumerate(enhanced_header, start=1):
            cell = txn_sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL

        excel_row = 2
        member_results = mtm_results.get('member_results', {})

        # Get code column index for XX detection
        code_col_idx = column_indices[0] if column_indices else 11

        for member_key, result in member_results.items():
            # Determine member status
            is_bp = False
            is_xx = False
            for txn in result.get('transactions', []):
                if self._is_bp_member(txn, column_indices, bp_config):
                    is_bp = True
                    break
                if self._is_xx_code(txn, code_col_idx):
                    is_xx = True

            has_flags = result.get('has_flags', False)
            if is_bp:
                status = "BP"
                fill = self.BP_FILL
            elif is_xx:
                status = "XX"
                fill = self.XX_FILL
            elif has_flags:
                status = "Flagged"
                fill = self.HIGHLIGHT_FILL
            else:
                status = "Valid"
                fill = None

            # Format flags
            flags = result.get('flags', [])
            flags_str = ' | '.join([str(flag) for flag in flags]) if flags else ''

            # Write each transaction
            for txn in result.get('transactions', []):
                row_data = txn + [status, flags_str]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = txn_sheet.cell(row=excel_row, column=col_idx, value=value)
                    if fill:
                        cell.fill = fill

                excel_row += 1

            self._apply_member_separator(txn_sheet, excel_row - 1, len(enhanced_header))

        # Auto-adjust widths
        self._auto_adjust_column_widths(txn_sheet)

    def create_all_types_report(
        self,
        header_row: List[str],
        type_results: Dict[str, Any],
        output_filename: str,
        column_mapping: Dict[str, int] = None,
        bp_config: Dict[str, Any] = None,
        member_type_mapping: Dict[str, str] = None
    ) -> str:
        """
        Create Excel report with tabs for each member_type plus a summary tab.

        Args:
            header_row: Column headers from original file
            type_results: Dictionary with results per member_type
            output_filename: Name for output file
            column_mapping: Dictionary with column name to index mappings
            bp_config: BP detection configuration
            member_type_mapping: Mapping of member_type codes to config keys

        Returns:
            Full path to generated report file
        """
        wb = Workbook()

        # Remove default sheet, we'll add our own
        wb.remove(wb.active)

        # Create Summary sheet first
        summary_sheet = wb.create_sheet("Summary", 0)
        self._add_all_types_summary_sheet(summary_sheet, type_results, member_type_mapping)

        # Get column indices for BP detection
        column_indices = []
        if column_mapping:
            bp_columns = bp_config.get('columns', ['code', 'member_type', 'member_group']) if bp_config else ['code', 'member_type', 'member_group']
            for col_name in bp_columns:
                idx = column_mapping.get(col_name)
                if idx is not None and idx >= 0:
                    column_indices.append(idx)
        if not column_indices:
            column_indices = [11, 9, 10]  # code, member_type, member_group in new format

        # Create a tab for each member_type
        for member_type, results in type_results.items():
            # Sanitize sheet name (max 31 chars, no special chars)
            sheet_name = member_type[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
            if not sheet_name:
                sheet_name = "Unknown"

            type_sheet = wb.create_sheet(sheet_name)

            if results.get('is_mtm'):
                # MTM type - use member-based format
                self._write_mtm_type_sheet(type_sheet, header_row, results, column_indices, bp_config)
            else:
                # Standard type or unknown - use row-based format
                self._write_standard_type_sheet(type_sheet, header_row, results, column_indices, bp_config)

        # Save workbook
        output_path = self.output_folder / output_filename
        wb.save(output_path)

        return str(output_path)

    def _add_all_types_summary_sheet(
        self,
        sheet,
        type_results: Dict[str, Any],
        member_type_mapping: Dict[str, str] = None
    ):
        """Add summary sheet for all types report"""

        # Title
        sheet['A1'] = "ALL MEMBERSHIP TYPES AUDIT SUMMARY"
        sheet['A1'].font = Font(bold=True, size=16)

        # Overall totals
        row = 3
        total_records = sum(r.get('total_records', 0) for r in type_results.values())
        total_flagged = 0
        total_financial_impact = 0

        for results in type_results.values():
            if results.get('is_mtm'):
                total_flagged += results.get('flagged_members', 0)
            else:
                total_flagged += results.get('flagged_count', 0)
            total_financial_impact += results.get('financial_impact', 0)

        sheet[f'A{row}'] = "Total Records:"
        sheet[f'B{row}'] = total_records
        sheet[f'A{row}'].font = self.BOLD_FONT

        row += 1
        sheet[f'A{row}'] = "Total Flagged:"
        sheet[f'B{row}'] = total_flagged
        sheet[f'A{row}'].font = self.BOLD_FONT
        sheet[f'B{row}'].fill = self.HIGHLIGHT_FILL

        row += 1
        flagged_pct = (total_flagged / total_records * 100) if total_records > 0 else 0
        sheet[f'A{row}'] = "Flagged Percentage:"
        sheet[f'B{row}'] = f"{flagged_pct:.1f}%"
        sheet[f'A{row}'].font = self.BOLD_FONT

        row += 1
        sheet[f'A{row}'] = "Total Financial Impact:"
        sheet[f'B{row}'] = f"${total_financial_impact:,.2f}"
        sheet[f'A{row}'].font = self.BOLD_FONT
        sheet[f'B{row}'].font = Font(bold=True, color="FF0000")

        # Breakdown by member type
        row += 3
        sheet[f'A{row}'] = "BREAKDOWN BY MEMBER TYPE"
        sheet[f'A{row}'].font = Font(bold=True, size=14)

        row += 2
        # Header row for breakdown table
        headers = ["Member Type", "Config", "Records", "Flagged", "Flag %", "Financial Impact", "Rules Applied"]
        for col_idx, header_text in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL

        row += 1

        # Sort by record count descending
        sorted_types = sorted(type_results.items(), key=lambda x: x[1].get('total_records', 0), reverse=True)

        for member_type, results in sorted_types:
            config_key = results.get('config_key', 'N/A')
            config_name = config_key if config_key else 'Unknown'

            records = results.get('total_records', 0)

            if results.get('is_mtm'):
                flagged = results.get('flagged_members', 0)
            else:
                flagged = results.get('flagged_count', 0)

            flag_pct = (flagged / records * 100) if records > 0 else 0
            financial_impact = results.get('financial_impact', 0)
            has_rules = "Yes" if results.get('has_rules') else "No (raw data)"

            sheet.cell(row=row, column=1, value=member_type)
            sheet.cell(row=row, column=2, value=config_name)
            sheet.cell(row=row, column=3, value=records)

            flagged_cell = sheet.cell(row=row, column=4, value=flagged)
            if flagged > 0:
                flagged_cell.fill = self.HIGHLIGHT_FILL

            sheet.cell(row=row, column=5, value=f"{flag_pct:.1f}%")
            sheet.cell(row=row, column=6, value=f"${financial_impact:,.2f}")
            sheet.cell(row=row, column=7, value=has_rules)

            row += 1

        # Legend
        row += 2
        sheet[f'A{row}'] = "LEGEND"
        sheet[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        sheet[f'A{row}'] = "Known Types:"
        sheet[f'B{row}'] = "1MCORE (1 Month PIF), 1YRCORE (1 Year PIF), 3MCORE (3 Month PIF), MTMCORE (Month to Month)"

        row += 1
        sheet[f'A{row}'] = "Unknown Types:"
        sheet[f'B{row}'] = "Raw data only - no red flag rules applied. Review data to define rules."

        row += 2
        yellow_cell = sheet.cell(row=row, column=1, value="Yellow")
        yellow_cell.fill = self.HIGHLIGHT_FILL
        sheet[f'B{row}'] = "= Flagged accounts (red flags detected)"

        row += 1
        blue_cell = sheet.cell(row=row, column=1, value="Light Blue")
        blue_cell.fill = self.XX_FILL
        sheet[f'B{row}'] = "= XX code accounts (code column = 'xx')"

        row += 1
        orange_cell = sheet.cell(row=row, column=1, value="Orange")
        orange_cell.fill = self.BP_FILL
        sheet[f'B{row}'] = "= Billing Problem accounts (BP indicator in code/member_type)"

        # Auto-adjust widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['G'].width = 18

    def _write_standard_type_sheet(
        self,
        sheet,
        header_row: List[str],
        results: Dict[str, Any],
        column_indices: List[int],
        bp_config: Dict[str, Any]
    ):
        """Write a sheet for a standard (non-MTM) member type"""

        # Check if this is a grouped result
        if results.get('is_grouped') and results.get('member_results'):
            self._write_grouped_type_sheet(sheet, header_row, results, column_indices, bp_config)
            return

        # Add Notes column to header
        enhanced_header = header_row + ["Notes"]
        col_count = len(enhanced_header)

        # Write header row
        for col_idx, header_text in enumerate(enhanced_header, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN

        audit_results = results.get('audit_results', [])
        rows = results.get('rows', [])

        # If we have audit_results with row_data, use those
        if audit_results and 'row_data' in audit_results[0]:
            data_pairs = [(r.get('row_data', []), r) for r in audit_results]
        else:
            # Pair rows with empty results
            data_pairs = [(row, {'red_flags': [], 'has_flags': False}) for row in rows]

        # Get code column index for XX detection
        code_col_idx = column_indices[0] if column_indices else 11  # code column is first in the list

        # Categorize rows
        flagged_rows, xx_rows, bp_rows, valid_rows = self._categorize_rows(
            [pair[0] for pair in data_pairs],
            [pair[1] for pair in data_pairs],
            column_indices,
            bp_config,
            code_col_idx
        )

        excel_row = 2

        # Section 1: FLAGGED ACCOUNTS
        if flagged_rows:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"FLAGGED ACCOUNTS ({len(flagged_rows)} records)",
                col_count
            )

            for data_row, audit_result in flagged_rows:
                red_flags = audit_result.get('red_flags', [])
                notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                enhanced_row = list(data_row) + [notes]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    cell = sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.HIGHLIGHT_FILL

                excel_row += 1

            excel_row += 1

        # Section 2: XX CODE ACCOUNTS
        if xx_rows:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"XX CODE ACCOUNTS ({len(xx_rows)} records)",
                col_count
            )

            for data_row, audit_result in xx_rows:
                red_flags = audit_result.get('red_flags', [])
                notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                enhanced_row = list(data_row) + [notes]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    cell = sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.XX_FILL

                excel_row += 1

            excel_row += 1

        # Section 3: BILLING PROBLEM ACCOUNTS
        if bp_rows:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"BILLING PROBLEM ACCOUNTS ({len(bp_rows)} records)",
                col_count
            )

            for data_row, audit_result in bp_rows:
                red_flags = audit_result.get('red_flags', [])
                notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                enhanced_row = list(data_row) + [notes]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    cell = sheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.fill = self.BP_FILL

                excel_row += 1

            excel_row += 1

        # Section 4: VALID ACCOUNTS
        if valid_rows:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"VALID ACCOUNTS ({len(valid_rows)} records)",
                col_count
            )

            for data_row, audit_result in valid_rows:
                enhanced_row = list(data_row) + [""]

                for col_idx, value in enumerate(enhanced_row, start=1):
                    sheet.cell(row=excel_row, column=col_idx, value=value)

                excel_row += 1

        # Auto-adjust column widths
        self._auto_adjust_column_widths(sheet)

    def _write_grouped_type_sheet(
        self,
        sheet,
        header_row: List[str],
        results: Dict[str, Any],
        column_indices: List[int],
        bp_config: Dict[str, Any]
    ):
        """Write a sheet for a grouped member type with Net Balance column"""

        # Add Net Balance and Notes columns to header
        enhanced_header = header_row + ["Net Balance", "Notes"]
        col_count = len(enhanced_header)

        # Write header row
        for col_idx, header_text in enumerate(enhanced_header, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN

        member_results = results.get('member_results', {})

        # Get code column index for XX detection
        code_col_idx = column_indices[0] if column_indices else 11

        # Categorize members into flagged, XX, BP, and valid
        flagged_members, xx_members, bp_members, valid_members = self._categorize_members(
            member_results, column_indices, bp_config, code_col_idx
        )

        excel_row = 2

        # Section 1: FLAGGED ACCOUNTS
        if flagged_members:
            total_flagged_rows = sum(len(m['transactions']) for m in flagged_members)
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"FLAGGED ACCOUNTS ({len(flagged_members)} members, {total_flagged_rows} records)",
                col_count
            )

            for member_result in flagged_members:
                flags = member_result.get('flags', [])
                notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        cell = sheet.cell(row=excel_row, column=col_idx, value=value)
                        cell.fill = self.HIGHLIGHT_FILL
                    excel_row += 1

                self._apply_member_separator(sheet, excel_row - 1, col_count)

            excel_row += 1

        # Section 2: XX CODE ACCOUNTS
        if xx_members:
            total_xx_rows = sum(len(m['transactions']) for m in xx_members)
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"XX CODE ACCOUNTS ({len(xx_members)} members, {total_xx_rows} records)",
                col_count
            )

            for member_result in xx_members:
                flags = member_result.get('flags', [])
                notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        cell = sheet.cell(row=excel_row, column=col_idx, value=value)
                        cell.fill = self.XX_FILL
                    excel_row += 1

                self._apply_member_separator(sheet, excel_row - 1, col_count)

            excel_row += 1

        # Section 3: BILLING PROBLEM ACCOUNTS
        if bp_members:
            total_bp_rows = sum(len(m['transactions']) for m in bp_members)
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"BILLING PROBLEM ACCOUNTS ({len(bp_members)} members, {total_bp_rows} records)",
                col_count
            )

            for member_result in bp_members:
                flags = member_result.get('flags', [])
                notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        cell = sheet.cell(row=excel_row, column=col_idx, value=value)
                        cell.fill = self.BP_FILL
                    excel_row += 1

                self._apply_member_separator(sheet, excel_row - 1, col_count)

            excel_row += 1

        # Section 4: VALID ACCOUNTS
        if valid_members:
            total_valid_rows = sum(len(m['transactions']) for m in valid_members)
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"VALID ACCOUNTS ({len(valid_members)} members, {total_valid_rows} records)",
                col_count
            )

            for member_result in valid_members:
                net_balance = member_result.get('net_balance', 0.0)

                for txn in member_result['transactions']:
                    enhanced_row = list(txn) + [f"${net_balance:.2f}", ""]
                    for col_idx, value in enumerate(enhanced_row, start=1):
                        sheet.cell(row=excel_row, column=col_idx, value=value)
                    excel_row += 1

                self._apply_member_separator(sheet, excel_row - 1, col_count)

        # Auto-adjust column widths
        self._auto_adjust_column_widths(sheet)

    def _write_mtm_type_sheet(
        self,
        sheet,
        header_row: List[str],
        results: Dict[str, Any],
        column_indices: List[int],
        bp_config: Dict[str, Any]
    ):
        """Write a sheet for MTM member type (member-based grouping)"""

        # Create member-level header (same as main MTM report)
        member_header = [
            "First Name", "Last Name", "Member #", "Join Date",
            "Member Type", "Enrollment Fee", "Initial Payment", "Coverage Start",
            "Months Paid", "Missing Months", "Annual Fee", "Unmatched Charges", "Notes"
        ]
        col_count = len(member_header)

        # Write header row
        for col_idx, header_text in enumerate(member_header, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN

        member_results = results.get('member_results', {})

        # Get code column index for XX detection
        code_col_idx = column_indices[0] if column_indices else 11

        # Categorize members
        flagged_members, xx_members, bp_members, valid_members = self._categorize_members(
            member_results, column_indices, bp_config, code_col_idx
        )

        excel_row = 2

        # Section 1: FLAGGED MEMBERS
        if flagged_members:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"FLAGGED MEMBERS ({len(flagged_members)} members)",
                col_count
            )

            for member_result in flagged_members:
                excel_row = self._write_mtm_member_row(sheet, excel_row, member_result, self.HIGHLIGHT_FILL)
                self._apply_member_separator(sheet, excel_row - 1, col_count)

            excel_row += 1

        # Section 2: XX CODE MEMBERS
        if xx_members:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"XX CODE MEMBERS ({len(xx_members)} members)",
                col_count
            )

            for member_result in xx_members:
                excel_row = self._write_mtm_member_row(sheet, excel_row, member_result, self.XX_FILL)
                self._apply_member_separator(sheet, excel_row - 1, col_count)

            excel_row += 1

        # Section 3: BILLING PROBLEM MEMBERS
        if bp_members:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"BILLING PROBLEM MEMBERS ({len(bp_members)} members)",
                col_count
            )

            for member_result in bp_members:
                excel_row = self._write_mtm_member_row(sheet, excel_row, member_result, self.BP_FILL)
                self._apply_member_separator(sheet, excel_row - 1, col_count)

            excel_row += 1

        # Section 4: VALID MEMBERS
        if valid_members:
            excel_row = self._write_section_header(
                sheet, excel_row,
                f"VALID MEMBERS ({len(valid_members)} members)",
                col_count
            )

            for member_result in valid_members:
                excel_row = self._write_mtm_member_row(sheet, excel_row, member_result, None)
                self._apply_member_separator(sheet, excel_row - 1, col_count)

        # Auto-adjust column widths
        self._auto_adjust_column_widths(sheet)

    def _build_mtm_row_data(self, result: Dict[str, Any]) -> list:
        """Build row data for MTM member with all new fields."""
        # Format dates
        join_date = result.get('join_date')
        join_date_str = join_date.strftime('%m/%d/%y') if join_date else ''

        coverage_start = result.get('coverage_start')
        coverage_start_str = coverage_start.strftime('%m/%d/%y') if coverage_start else ''

        # Format missing months
        missing_months = result.get('missing_months', [])
        missing_str = ', '.join(missing_months[:5]) if missing_months else ''
        if len(missing_months) > 5:
            missing_str += f' (+{len(missing_months) - 5} more)'

        # Format enrollment fee status
        has_enrollment = result.get('has_enrollment_fee', False)
        enrollment_str = 'Yes' if has_enrollment else 'No'

        # Format initial payment
        has_initial = result.get('has_initial_payment', False)
        initial_amount = result.get('initial_payment_amount')
        if has_initial and initial_amount is not None:
            initial_str = f'Yes (${abs(initial_amount):.2f})'
        else:
            initial_str = 'No'

        # Format annual fee status
        has_annual = result.get('has_annual_fee', False)
        annual_str = 'Yes' if has_annual else 'No'

        # Unmatched charges count
        unmatched_count = result.get('unmatched_charges', 0)
        unmatched_str = str(unmatched_count) if unmatched_count > 0 else ''

        # Format notes from flags
        flags = result.get('flags', [])
        notes = ' | '.join([str(flag) for flag in flags]) if flags else ''

        return [
            result.get('first_name', ''),
            result.get('last_name', ''),
            result.get('member_number', ''),
            join_date_str,
            result.get('member_type', 'Unknown'),
            enrollment_str,
            initial_str,
            coverage_start_str,
            result.get('months_paid_count', 0),
            missing_str,
            annual_str,
            unmatched_str,
            notes
        ]

    def _write_mtm_member_row(self, sheet, row_num: int, member_result: Dict[str, Any], fill) -> int:
        """Write a single MTM member row and return next row number"""
        row_data = self._build_mtm_row_data(member_result)

        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_idx, value=value)
            if fill:
                cell.fill = fill

        return row_num + 1

    def create_individual_type_files(
        self,
        header_row: List[str],
        type_results: Dict[str, Any],
        base_filename: str,
        member_type_mapping: Dict[str, str],
        column_mapping: Dict[str, int] = None,
        bp_config: Dict[str, Any] = None
    ) -> Dict[str, Dict[str, str]]:
        """
        Create separate Excel files for each member_type.

        For known types (1MCORE, 1YRCORE, 3MCORE, MTMCORE):
            - Audited file with Notes column, organized as:
              Flagged (yellow) → Valid → BP (orange at bottom)
            - Raw file without Notes (for re-verification)
        For unknown types: Raw data only (original columns, no Notes)

        Args:
            header_row: Column headers from original file
            type_results: Dictionary with results per member_type
            base_filename: Base name for output files (without extension)
            member_type_mapping: Mapping of member_type codes to config keys
            column_mapping: Dictionary with column name to index mappings
            bp_config: BP detection configuration

        Returns:
            Dict mapping member_type to dict with 'audited' and/or 'raw' file paths
        """
        individual_files = {}

        # Get column indices for BP detection
        column_indices = []
        if column_mapping:
            bp_columns = bp_config.get('columns', ['code', 'member_type', 'member_group']) if bp_config else ['code', 'member_type', 'member_group']
            for col_name in bp_columns:
                idx = column_mapping.get(col_name)
                if idx is not None and idx >= 0:
                    column_indices.append(idx)
        if not column_indices:
            column_indices = [11, 9, 10]  # code, member_type, member_group in new format

        for member_type, results in type_results.items():
            is_known_type = results.get('has_rules', False)
            rows = results.get('rows', [])

            if not rows:
                continue

            # Sanitize member_type for filename
            safe_type = member_type.replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')

            individual_files[member_type] = {}

            if is_known_type:
                # Known type: generate BOTH audited file (with Notes) AND raw file (for re-upload)
                is_grouped = results.get('is_grouped', False)

                # 1. Create audited file with grouped or row-based format
                wb_audited = Workbook()
                sheet_audited = wb_audited.active
                sheet_audited.title = "Data"

                if is_grouped and results.get('member_results'):
                    # Grouped format: Net Balance + Notes columns
                    enhanced_header = header_row + ["Net Balance", "Notes"]
                    col_count = len(enhanced_header)

                    for col_idx, header_text in enumerate(enhanced_header, start=1):
                        cell = sheet_audited.cell(row=1, column=col_idx, value=header_text)
                        cell.font = self.BOLD_FONT
                        cell.fill = self.HEADER_FILL
                        cell.alignment = self.CENTER_ALIGN

                    member_results = results.get('member_results', {})

                    # Get code column index for XX detection
                    code_col_idx = column_indices[0] if column_indices else 11

                    # Categorize members
                    flagged_members, xx_members_list, bp_members_list, valid_members_list = self._categorize_members(
                        member_results, column_indices, bp_config, code_col_idx
                    )

                    excel_row = 2

                    # Section 1: FLAGGED ACCOUNTS (yellow)
                    if flagged_members:
                        total_flagged_rows = sum(len(m['transactions']) for m in flagged_members)
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"FLAGGED ACCOUNTS ({len(flagged_members)} members, {total_flagged_rows} records)",
                            col_count
                        )
                        for member_result in flagged_members:
                            flags = member_result.get('flags', [])
                            notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                            net_balance = member_result.get('net_balance', 0.0)
                            for txn in member_result['transactions']:
                                enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                                for col_idx, value in enumerate(enhanced_row, start=1):
                                    cell = sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                                    cell.fill = self.HIGHLIGHT_FILL
                                excel_row += 1
                            self._apply_member_separator(sheet_audited, excel_row - 1, col_count)
                        excel_row += 1

                    # Section 2: XX CODE ACCOUNTS (light blue)
                    if xx_members_list:
                        total_xx_rows = sum(len(m['transactions']) for m in xx_members_list)
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"XX CODE ACCOUNTS ({len(xx_members_list)} members, {total_xx_rows} records)",
                            col_count
                        )
                        for member_result in xx_members_list:
                            flags = member_result.get('flags', [])
                            notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                            net_balance = member_result.get('net_balance', 0.0)
                            for txn in member_result['transactions']:
                                enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                                for col_idx, value in enumerate(enhanced_row, start=1):
                                    cell = sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                                    cell.fill = self.XX_FILL
                                excel_row += 1
                            self._apply_member_separator(sheet_audited, excel_row - 1, col_count)
                        excel_row += 1

                    # Section 3: VALID ACCOUNTS (no highlight)
                    if valid_members_list:
                        total_valid_rows = sum(len(m['transactions']) for m in valid_members_list)
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"VALID ACCOUNTS ({len(valid_members_list)} members, {total_valid_rows} records)",
                            col_count
                        )
                        for member_result in valid_members_list:
                            net_balance = member_result.get('net_balance', 0.0)
                            for txn in member_result['transactions']:
                                enhanced_row = list(txn) + [f"${net_balance:.2f}", ""]
                                for col_idx, value in enumerate(enhanced_row, start=1):
                                    sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                                excel_row += 1
                            self._apply_member_separator(sheet_audited, excel_row - 1, col_count)
                        excel_row += 1

                    # Section 4: BILLING PROBLEM ACCOUNTS (orange at bottom)
                    if bp_members_list:
                        total_bp_rows = sum(len(m['transactions']) for m in bp_members_list)
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"BILLING PROBLEM ACCOUNTS ({len(bp_members_list)} members, {total_bp_rows} records)",
                            col_count
                        )
                        for member_result in bp_members_list:
                            flags = member_result.get('flags', [])
                            notes = " | ".join([str(flag) for flag in flags]) if flags else ""
                            net_balance = member_result.get('net_balance', 0.0)
                            for txn in member_result['transactions']:
                                enhanced_row = list(txn) + [f"${net_balance:.2f}", notes]
                                for col_idx, value in enumerate(enhanced_row, start=1):
                                    cell = sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                                    cell.fill = self.BP_FILL
                                excel_row += 1
                            self._apply_member_separator(sheet_audited, excel_row - 1, col_count)

                else:
                    # Non-grouped format: Notes column only
                    enhanced_header = header_row + ["Notes"]
                    col_count = len(enhanced_header)

                    for col_idx, header_text in enumerate(enhanced_header, start=1):
                        cell = sheet_audited.cell(row=1, column=col_idx, value=header_text)
                        cell.font = self.BOLD_FONT
                        cell.fill = self.HEADER_FILL
                        cell.alignment = self.CENTER_ALIGN

                    audit_results = results.get('audit_results', [])
                    if audit_results and len(audit_results) > 0:
                        data_pairs = list(zip(rows, audit_results))
                    else:
                        data_pairs = [(row, {'red_flags': [], 'has_flags': False}) for row in rows]

                    code_col_idx = column_indices[0] if column_indices else 11

                    flagged_rows, xx_rows, bp_rows, valid_rows = self._categorize_rows(
                        [pair[0] for pair in data_pairs],
                        [pair[1] for pair in data_pairs],
                        column_indices,
                        bp_config,
                        code_col_idx
                    )

                    excel_row = 2

                    if flagged_rows:
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"FLAGGED ACCOUNTS ({len(flagged_rows)} records)",
                            col_count
                        )
                        for data_row, audit_result in flagged_rows:
                            red_flags = audit_result.get('red_flags', [])
                            notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                            enhanced_row = list(data_row) + [notes]
                            for col_idx, value in enumerate(enhanced_row, start=1):
                                cell = sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                                cell.fill = self.HIGHLIGHT_FILL
                            excel_row += 1
                        excel_row += 1

                    if xx_rows:
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"XX CODE ACCOUNTS ({len(xx_rows)} records)",
                            col_count
                        )
                        for data_row, audit_result in xx_rows:
                            red_flags = audit_result.get('red_flags', [])
                            notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                            enhanced_row = list(data_row) + [notes]
                            for col_idx, value in enumerate(enhanced_row, start=1):
                                cell = sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                                cell.fill = self.XX_FILL
                            excel_row += 1
                        excel_row += 1

                    if valid_rows:
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"VALID ACCOUNTS ({len(valid_rows)} records)",
                            col_count
                        )
                        for data_row, audit_result in valid_rows:
                            enhanced_row = list(data_row) + [""]
                            for col_idx, value in enumerate(enhanced_row, start=1):
                                sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                            excel_row += 1
                        excel_row += 1

                    if bp_rows:
                        excel_row = self._write_section_header(
                            sheet_audited, excel_row,
                            f"BILLING PROBLEM ACCOUNTS ({len(bp_rows)} records)",
                            col_count
                        )
                        for data_row, audit_result in bp_rows:
                            red_flags = audit_result.get('red_flags', [])
                            notes = " | ".join([str(flag) for flag in red_flags]) if red_flags else ""
                            enhanced_row = list(data_row) + [notes]
                            for col_idx, value in enumerate(enhanced_row, start=1):
                                cell = sheet_audited.cell(row=excel_row, column=col_idx, value=value)
                                cell.fill = self.BP_FILL
                            excel_row += 1

                self._auto_adjust_column_widths(sheet_audited)

                audited_filename = f"{base_filename}_{safe_type}.xlsx"
                audited_path = self.output_folder / audited_filename
                wb_audited.save(audited_path)
                individual_files[member_type]['audited'] = str(audited_path)

                # 2. Create raw file without Notes (for re-verification)
                wb_raw = Workbook()
                sheet_raw = wb_raw.active
                sheet_raw.title = "Data"

                # Write header row (no Notes)
                for col_idx, header_text in enumerate(header_row, start=1):
                    cell = sheet_raw.cell(row=1, column=col_idx, value=header_text)
                    cell.font = self.BOLD_FONT
                    cell.fill = self.HEADER_FILL
                    cell.alignment = self.CENTER_ALIGN

                # Write data rows
                excel_row = 2
                for row in rows:
                    for col_idx, value in enumerate(row, start=1):
                        sheet_raw.cell(row=excel_row, column=col_idx, value=value)
                    excel_row += 1

                self._auto_adjust_column_widths(sheet_raw)

                raw_filename = f"{base_filename}_{safe_type}_raw.xlsx"
                raw_path = self.output_folder / raw_filename
                wb_raw.save(raw_path)
                individual_files[member_type]['raw'] = str(raw_path)

            else:
                # Unknown type: raw data only, no Notes column
                wb = Workbook()
                sheet = wb.active
                sheet.title = "Data"

                # Write header row
                for col_idx, header_text in enumerate(header_row, start=1):
                    cell = sheet.cell(row=1, column=col_idx, value=header_text)
                    cell.font = self.BOLD_FONT
                    cell.fill = self.HEADER_FILL
                    cell.alignment = self.CENTER_ALIGN

                # Write data rows
                excel_row = 2
                for row in rows:
                    for col_idx, value in enumerate(row, start=1):
                        sheet.cell(row=excel_row, column=col_idx, value=value)
                    excel_row += 1

                self._auto_adjust_column_widths(sheet)

                output_filename = f"{base_filename}_{safe_type}.xlsx"
                output_path = self.output_folder / output_filename
                wb.save(output_path)
                individual_files[member_type]['raw'] = str(output_path)

        return individual_files

    def create_split_type_files(
        self,
        header_row: List[str],
        rows_by_type: Dict[str, List[List[str]]],
        base_filename: str
    ) -> Dict[str, Dict[str, Any]]:
        """
        Create separate raw Excel files for each member_type.
        NO rules applied, NO highlighting, NO Notes column - pure raw data output.

        Args:
            header_row: Column headers from original file
            rows_by_type: Dictionary mapping member_type to list of rows
            base_filename: Base name for output files (without extension)

        Returns:
            Dict mapping member_type to dict with 'file_path', 'row_count', 'file_size'
        """
        split_files = {}

        for member_type, rows in rows_by_type.items():
            if not rows:
                continue

            # Sanitize member_type for filename
            safe_type = member_type.replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '').replace(':', '-')
            if not safe_type:
                safe_type = 'UNKNOWN'

            # Create workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Data"

            # Write header row with basic formatting
            for col_idx, header_text in enumerate(header_row, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = self.BOLD_FONT
                cell.fill = self.HEADER_FILL

            # Write data rows - no highlighting, no modifications
            excel_row = 2
            for row in rows:
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=excel_row, column=col_idx, value=value)
                excel_row += 1

            # Auto-adjust column widths
            self._auto_adjust_column_widths(sheet)

            # Generate output filename
            output_filename = f"{base_filename}_{safe_type}.xlsx"
            output_path = self.output_folder / output_filename
            wb.save(output_path)

            # Get file size
            file_size = output_path.stat().st_size

            split_files[member_type] = {
                'file_path': str(output_path),
                'row_count': len(rows),
                'file_size': file_size,
                'filename': output_filename
            }

        return split_files
